"""
================================================================================
 三层递进漏斗推荐引擎 v6.0 — Funnel Engine
================================================================================
 核心算法：
   Layer 1：学科门类初筛 — 认知风格 + 人格倾向 → 13 个门类匹配分
   Layer 2：专业类精选 — 产业 + 资产 + 分数 → Top 8 硬截断
   Layer 3：专业微观狙击 — 微观动作 + 热度 + 红线 → ≤6/类硬截断

  Layer 3 公式（用户确认）：
   score = (micro_match × 0.6 + heat_align × 0.4) × threshold_pass
   threshold_pass ∈ {0, 1} → 红线触发则直接归零
================================================================================
"""

import json
import math
import os
from pathlib import Path
from typing import Optional
from dataclasses import dataclass, field

import openpyxl

from user_profile import (
    UserProfile,
    BEHAVIOR_DIMENSIONS,
    INDUSTRY_CLUSTERS,
)

# ============================================================================
# 路径配置
# ============================================================================
BASE_DIR = Path(__file__).parent
CHECKPOINT_DIR = BASE_DIR / "label_checkpoints"
MAJORS_XLSX = BASE_DIR / "gaokao_majors.xlsx"

# ============================================================================
# 标签 → 向量映射表（将分类标签转为数值向量用于余弦匹配）
# ============================================================================

# 认知风格 → 10 维行为向量
COGNITIVE_STYLE_TO_BEHAVIOR: dict[str, dict[str, float]] = {
    "系统建构":    {"逻辑推理": 0.60, "精细操作": 0.40},
    "逻辑推演":    {"逻辑推理": 0.70, "数据敏感": 0.30},
    "空间想象":    {"创造性思维": 0.50, "动手实验": 0.50},
    "语言敏感":    {"沟通表达": 0.60, "记忆积累": 0.40},
    "抽象思辨":    {"逻辑推理": 0.50, "创造性思维": 0.50},
    "共情表达":    {"沟通表达": 0.60, "团队协作": 0.40},
    "数据敏感":    {"数据敏感": 0.60, "逻辑推理": 0.40},
    "实验思维":    {"动手实验": 0.60, "逻辑推理": 0.40},
    "归纳推理":    {"逻辑推理": 0.50, "数据敏感": 0.50},
    "动手实践":    {"动手实验": 0.60, "精细操作": 0.40},
    "审美判断":    {"创造性思维": 0.60, "精细操作": 0.40},
    "社会洞察":    {"沟通表达": 0.50, "团队协作": 0.50},
    "批判思维":    {"逻辑推理": 0.60, "创造性思维": 0.40},
    "记忆积累":    {"记忆积累": 0.70, "持续专注": 0.30},
    "模式识别":    {"数据敏感": 0.50, "逻辑推理": 0.50},
    "流程管控":    {"精细操作": 0.60, "团队协作": 0.40},
    "创意发散":    {"创造性思维": 0.70, "沟通表达": 0.30},
}

# 人格倾向 → RIASEC 六维向量
PERSONA_TO_RIASEC: dict[str, dict[str, float]] = {
    "理性分析型":  {"I": 0.70, "C": 0.30},
    "动手实践型":  {"R": 0.70, "I": 0.30},
    "情感输出型":  {"A": 0.70, "S": 0.30},
    "审美直觉型":  {"A": 0.80, "I": 0.20},
    "社会服务型":  {"S": 0.80, "E": 0.20},
    "探索创新型":  {"I": 0.50, "A": 0.50},
    "规则执行型":  {"C": 0.80, "R": 0.20},
    "领导管理型":  {"E": 0.70, "S": 0.30},
    "沟通协作型":  {"S": 0.50, "E": 0.50},
    "数据驱动型":  {"I": 0.40, "C": 0.60},
}

# 微观动作 → 10 维行为向量
MICRO_ACTION_TO_BEHAVIOR: dict[str, dict[str, float]] = {
    "Debug与迭代修复":          {"逻辑推理": 0.50, "动手实验": 0.30, "持续专注": 0.20},
    "系统架构设计":             {"逻辑推理": 0.50, "创造性思维": 0.30, "精细操作": 0.20},
    "代码编写与Review":         {"逻辑推理": 0.40, "精细操作": 0.40, "团队协作": 0.20},
    "数据分析与统计建模":       {"数据敏感": 0.50, "逻辑推理": 0.50},
    "机器学习模型训练":         {"数据敏感": 0.40, "逻辑推理": 0.30, "动手实验": 0.30},
    "数据库设计与优化":         {"逻辑推理": 0.40, "精细操作": 0.40, "数据敏感": 0.20},
    "实验设计与执行":           {"动手实验": 0.50, "逻辑推理": 0.30, "精细操作": 0.20},
    "文献检索与综述撰写":       {"记忆积累": 0.40, "逻辑推理": 0.30, "持续专注": 0.30},
    "田野调查与访谈":           {"沟通表达": 0.50, "团队协作": 0.30, "抗压能力": 0.20},
    "临床诊断与鉴别":           {"逻辑推理": 0.50, "记忆积累": 0.30, "精细操作": 0.20},
    "手术/介入操作":            {"动手实验": 0.50, "精细操作": 0.30, "抗压能力": 0.20},
    "医学影像判读":             {"数据敏感": 0.40, "精细操作": 0.30, "记忆积累": 0.30},
    "法律文书起草与审查":       {"逻辑推理": 0.40, "沟通表达": 0.30, "记忆积累": 0.30},
    "庭审辩论与质证":           {"沟通表达": 0.50, "逻辑推理": 0.30, "抗压能力": 0.20},
    "合同谈判与尽调":           {"沟通表达": 0.40, "数据敏感": 0.30, "逻辑推理": 0.30},
    "财务报表编制与分析":       {"数据敏感": 0.50, "精细操作": 0.30, "逻辑推理": 0.20},
    "风险评估与量化":           {"数据敏感": 0.50, "逻辑推理": 0.50},
    "审计底稿编制":             {"精细操作": 0.50, "数据敏感": 0.30, "记忆积累": 0.20},
    "教学设计/教案编写":        {"创造性思维": 0.40, "沟通表达": 0.30, "记忆积累": 0.30},
    "课堂讲授与互动":           {"沟通表达": 0.60, "团队协作": 0.20, "创造性思维": 0.20},
    "课堂教学与互动":           {"沟通表达": 0.60, "团队协作": 0.20, "创造性思维": 0.20},
    "学生心理辅导":             {"沟通表达": 0.50, "团队协作": 0.30, "记忆积累": 0.20},
    "素描/色彩/造型创作":       {"创造性思维": 0.50, "动手实验": 0.30, "精细操作": 0.20},
    "软件UI/UX设计":            {"创造性思维": 0.40, "精细操作": 0.30, "沟通表达": 0.30},
    "三维建模与渲染":           {"动手实验": 0.40, "创造性思维": 0.30, "精细操作": 0.30},
    "乐器演奏/声乐训练":        {"动手实验": 0.40, "精细操作": 0.30, "记忆积累": 0.30},
    "剧本分析与表演创作":       {"创造性思维": 0.40, "沟通表达": 0.40, "记忆积累": 0.20},
    "镜头语言与剪辑":           {"创造性思维": 0.50, "精细操作": 0.30, "数据敏感": 0.20},
    "体能训练与运动康复":       {"动手实验": 0.50, "持续专注": 0.30, "记忆积累": 0.20},
    "竞赛战术制定":             {"逻辑推理": 0.40, "数据敏感": 0.30, "团队协作": 0.30},
    "运动生物力学分析":         {"数据敏感": 0.40, "逻辑推理": 0.30, "动手实验": 0.30},
    "工程制图/CAD建模":         {"动手实验": 0.40, "精细操作": 0.30, "创造性思维": 0.30},
    "电路设计与PCB布局":        {"逻辑推理": 0.40, "动手实验": 0.30, "精细操作": 0.30},
    "嵌入式系统开发":           {"逻辑推理": 0.40, "动手实验": 0.40, "精细操作": 0.20},
    "化学合成与分离纯化":       {"动手实验": 0.50, "精细操作": 0.30, "持续专注": 0.20},
    "色谱/光谱分析":            {"数据敏感": 0.40, "精细操作": 0.30, "逻辑推理": 0.30},
    "材料性能测试":             {"动手实验": 0.40, "数据敏感": 0.30, "精细操作": 0.30},
    "环境采样与监测":           {"动手实验": 0.40, "数据敏感": 0.30, "持续专注": 0.30},
    "环评报告编制":             {"逻辑推理": 0.30, "沟通表达": 0.30, "记忆积累": 0.40},
    "生态修复方案设计":         {"创造性思维": 0.40, "逻辑推理": 0.30, "动手实验": 0.30},
    "政策文本分析与解读":       {"逻辑推理": 0.50, "沟通表达": 0.30, "记忆积累": 0.20},
    "新闻采访与稿件撰写":       {"沟通表达": 0.50, "记忆积累": 0.30, "创造性思维": 0.20},
    "多语种翻译与本地化":       {"记忆积累": 0.50, "沟通表达": 0.30, "精细操作": 0.20},
    "用户需求调研与访谈":       {"沟通表达": 0.50, "团队协作": 0.30, "数据敏感": 0.20},
    "产品原型设计与迭代":       {"创造性思维": 0.50, "动手实验": 0.30, "团队协作": 0.20},
    "A/B测试与增长实验":        {"数据敏感": 0.50, "逻辑推理": 0.30, "创造性思维": 0.20},
    "芯片版图设计":             {"精细操作": 0.40, "逻辑推理": 0.30, "动手实验": 0.30},
    "信号完整性分析":           {"逻辑推理": 0.50, "数据敏感": 0.30, "精细操作": 0.20},
    "射频调试与匹配":           {"动手实验": 0.50, "逻辑推理": 0.30, "数据敏感": 0.20},
    "工地现场管理与监理":       {"团队协作": 0.40, "抗压能力": 0.30, "精细操作": 0.30},
    "结构力学计算与分析":       {"逻辑推理": 0.50, "数据敏感": 0.50},
    "造价预算编制":             {"数据敏感": 0.50, "精细操作": 0.30, "记忆积累": 0.20},
    "临床护理操作":             {"动手实验": 0.40, "精细操作": 0.30, "沟通表达": 0.30},
    "康复训练方案制定":         {"创造性思维": 0.40, "沟通表达": 0.30, "动手实验": 0.30},
    "公共卫生流行病学调查":     {"数据敏感": 0.40, "逻辑推理": 0.30, "团队协作": 0.30},
}

# 新产业标签 → 旧 10 产业集群映射（用于向后兼容现有问卷）
INDUSTRY_TAG_TO_CLUSTER: dict[str, str] = {
    "互联网/软件":       "互联网与软件",
    "人工智能/大模型":    "AI与大模型",
    "半导体/集成电路":    "半导体与芯片",
    "通信/5G/6G":        "互联网与软件",
    "智能制造/机器人":    "智能制造",
    "新能源汽车":         "新能源",
    "金融/银行":          "金融科技",
    "证券/基金/资管":     "金融科技",
    "保险/精算":          "金融科技",
    "医疗健康/临床":      "生物医药",
    "制药/生物技术":      "生物医药",
    "教育培训":           "教育培训",
    "法律服务/合规":      "政府公共",
    "建筑/土木/城规":     "智能制造",
    "能源/电力/碳中和":   "新能源",
    "石油/化工/材料":     "新能源",
    "农业/食品/林业":     "生物医药",
    "传媒/广告/公关":     "文化传媒",
    "影视/动画/娱乐":     "文化传媒",
    "艺术设计/文创":      "文化传媒",
    "体育产业":           "教育培训",
    "政府/公共服务":      "政府公共",
    "军事/国防工业":      "政府公共",
    "科研/学术":          "AI与大模型",
    "环保/新能源":        "新能源",
    "物流/供应链":        "智能制造",
    "零售/电商/消费":     "互联网与软件",
    "房地产/物业":        "政府公共",
    "咨询/审计/税务":     "金融科技",
    "旅游/酒店/会展":     "文化传媒",
    "航空航天":           "智能制造",
    "海洋工程/船舶":      "智能制造",
    "测绘/地质/矿业":     "新能源",
}

# 热度匹配：风险容忍度 → 社会热度的匹配分
# 行：风险容忍度等级，列：社会热度等级
HEAT_ALIGNMENT_MATRIX: dict[str, dict[str, float]] = {
    # 高风险容忍 → 喜欢追风口，极高/高热 = 高分
    "high":    {"极高": 1.0, "高": 0.9, "中": 0.5, "低": 0.2},
    # 中风险容忍 → 适中，匹配高/中 = 高分
    "medium":  {"极高": 0.6, "高": 0.8, "中": 1.0, "低": 0.5},
    # 低风险容忍 → 追求稳定，低/中热 = 高分
    "low":     {"极高": 0.2, "高": 0.4, "中": 0.8, "低": 1.0},
}

# 资产敏感度匹配：家庭经济水平 → 资产敏感度的匹配分
# 行：经济水平，列：资产敏感度
ASSET_ALIGNMENT_MATRIX: dict[str, dict[str, float]] = {
    # 高经济水平 → 高敏感也能驾驭
    "高":  {"低": 0.7, "中": 0.9, "高": 1.0},
    # 中经济水平 → 低/中敏感合适
    "中":  {"低": 1.0, "中": 0.9, "高": 0.5},
    # 低经济水平 → 低敏感最友好
    "低":  {"低": 1.0, "中": 0.6, "高": 0.2},
}

# 分数敏感度匹配：排名百分位 → 分数敏感度的匹配分
# 行：排名分位，列：分数敏感度
SCORE_ALIGNMENT_MATRIX: dict[str, dict[str, float]] = {
    # 前10% → 顶级排名，极高敏感也能打
    "top":       {"极高": 1.0, "高": 1.0, "中": 0.6, "低": 0.3},
    # 前10-30% → 中上排名
    "upper":     {"极高": 0.6, "高": 0.9, "中": 0.9, "低": 0.4},
    # 前30-60% → 中等排名
    "middle":    {"极高": 0.3, "高": 0.5, "中": 0.9, "低": 0.8},
    # 前60-100% → 中低排名
    "lower":     {"极高": 0.1, "高": 0.3, "中": 0.6, "低": 1.0},
}


def _rank_tier(percentile: float) -> str:
    if percentile <= 10:
        return "top"
    elif percentile <= 30:
        return "upper"
    elif percentile <= 60:
        return "middle"
    else:
        return "lower"


def _risk_tier(risk_tolerance: float) -> str:
    if risk_tolerance >= 70:
        return "high"
    elif risk_tolerance >= 40:
        return "medium"
    else:
        return "low"


# 招生体量 → 市场容量系数
ENROLLMENT_CAPACITY_COEFFICIENT: dict[str, float] = {
    "极大": 1.15,
    "大":   1.05,
    "中":   1.00,
    "小":   0.85,
    "极小": 0.70,
}

# 产业热度微调系数（用于 Layer 2 专业类级别）
INDUSTRY_HEAT_BONUS: dict[str, float] = {
    "人工智能/大模型":    1.05,
    "半导体/集成电路":    1.05,
    "新能源汽车":         1.04,
    "智能制造/机器人":    1.03,
    "新能源":            1.03,
    "互联网/软件":       1.02,
    "金融/银行":         1.02,
    "医疗健康/临床":     1.02,
    "航空航天":          1.04,
}


# ============================================================================
# 工具函数
# ============================================================================

def cosine_similarity(a: list[float], b: list[float]) -> float:
    """余弦相似度，钳位到 [0, 1]"""
    if len(a) != len(b) or len(a) == 0:
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(y * y for y in b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return max(0.0, min(1.0, dot / (norm_a * norm_b)))


def tags_to_vector(
    tags: list[str],
    mapping: dict[str, dict[str, float]],
    dims: list[str],
) -> list[float]:
    """将标签列表通过映射表转换为归一化向量"""
    vec = {d: 0.0 for d in dims}
    if not tags:
        return [0.5] * len(dims)  # 无标签时返回中性向量
    for tag in tags:
        if tag in mapping:
            for dim, weight in mapping[tag].items():
                if dim in vec:
                    vec[dim] += weight
    # 归一化：除以标签数量
    n = len(tags)
    if n > 0:
        vec = {k: min(1.0, v / n) for k, v in vec.items()}
    return [vec[d] for d in dims]


# ============================================================================
# 数据加载
# ============================================================================

@dataclass
class FunnelData:
    """三层漏斗数据结构"""
    disciplines: dict[str, dict] = field(default_factory=dict)
    categories: dict[str, dict] = field(default_factory=dict)
    majors: list[dict] = field(default_factory=list)
    # 索引
    _cat_to_disc: dict[str, str] = field(default_factory=dict)
    _cat_to_majors: dict[str, list[dict]] = field(default_factory=dict)

    def get_majors_in_category(self, category: str) -> list[dict]:
        return self._cat_to_majors.get(category, [])

    def get_discipline(self, category: str) -> str:
        return self._cat_to_disc.get(category, "")


def load_funnel_data() -> FunnelData:
    """加载三层标签数据"""
    data = FunnelData()
    errors = []

    # Layer 1: 学科门类
    l1_path = CHECKPOINT_DIR / "layer1_disciplines.json"
    if l1_path.exists():
        with open(l1_path, "r", encoding="utf-8") as f:
            data.disciplines = json.load(f)
    else:
        errors.append(f"L1 missing: {l1_path}")

    # Layer 2: 专业类
    l2_path = CHECKPOINT_DIR / "layer2_categories.json"
    if l2_path.exists():
        with open(l2_path, "r", encoding="utf-8") as f:
            data.categories = json.load(f)
    else:
        errors.append(f"L2 missing: {l2_path}")

    # Layer 3: 从 JSON 加载标签
    l3_path = CHECKPOINT_DIR / "layer3_majors.json"
    l3_data = {}
    if l3_path.exists():
        with open(l3_path, "r", encoding="utf-8") as f:
            l3_data = json.load(f)
    else:
        errors.append(f"L3 missing: {l3_path}")

    # 从 Excel 读取专业元数据
    if not MAJORS_XLSX.exists():
        errors.append(f"Excel missing: {MAJORS_XLSX}")
    else:
        wb = openpyxl.load_workbook(MAJORS_XLSX)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            seq, discipline, category, code, name = row
            if not category or str(category).strip() in ("", "-"):
                category = "交叉类"
            category = str(category).strip()
            discipline = str(discipline).strip() if discipline else ""
            code = str(code).strip() if code else ""
            name = str(name).strip() if name else ""

            major_entry = {
                "seq": seq,
                "discipline": discipline,
                "category": category,
                "code": code,
                "name": name,
                "micro_actions": l3_data.get(name, {}).get("micro_actions", []),
                "hard_threshold": l3_data.get(name, {}).get("hard_threshold", []),
                "social_heat": l3_data.get(name, {}).get("social_heat", "中"),
                "heat_trend": l3_data.get(name, {}).get("heat_trend", "平稳"),
                "enrollment_volume": l3_data.get(name, {}).get("enrollment_volume", "中"),
            }
            data.majors.append(major_entry)
            data._cat_to_disc[category] = discipline
            if category not in data._cat_to_majors:
                data._cat_to_majors[category] = []
            data._cat_to_majors[category].append(major_entry)

    # 校验
    data._load_errors = errors
    if not data.disciplines:
        print(f"[WARN] load_funnel_data: 0 disciplines loaded!")
    if not data.categories:
        print(f"[WARN] load_funnel_data: 0 categories loaded!")
    if not data.majors:
        print(f"[WARN] load_funnel_data: 0 majors loaded!")
    if errors:
        print(f"[WARN] load_funnel_data errors: {errors}")

    return data


# ============================================================================
# Layer 1: 学科门类匹配
# ============================================================================

def layer1_discipline_match(user: UserProfile, data: FunnelData) -> list[dict]:
    """
    第一层：学科门类初筛
    匹配用户的认知风格 + 人格倾向 → 13 个门类得分

    score = cosine(behavior, disc_cognitive) × 0.5
          + cosine(personality, disc_persona) × 0.3
          + discipline_weight × 0.2
    """
    # 用户向量
    user_behavior = [user.micro_behavior_vector.get(d, 50.0) / 100.0 for d in BEHAVIOR_DIMENSIONS]
    riasec_dims = ["R", "I", "A", "S", "E", "C"]
    user_personality = [user.inferred_personality.get(d, 50.0) / 100.0 for d in riasec_dims]

    results = []
    for disc_name, disc_labels in data.disciplines.items():
        cognitive_tags = disc_labels.get("cognitive_style", [])
        persona_tags = disc_labels.get("persona_tendency", [])
        disc_weight = disc_labels.get("discipline_weight", 0.5)

        # 标签 → 向量
        disc_cognitive_vec = tags_to_vector(cognitive_tags, COGNITIVE_STYLE_TO_BEHAVIOR, BEHAVIOR_DIMENSIONS)
        disc_persona_vec = tags_to_vector(persona_tags, PERSONA_TO_RIASEC, riasec_dims)

        cog_sim = cosine_similarity(user_behavior, disc_cognitive_vec)
        per_sim = cosine_similarity(user_personality, disc_persona_vec)

        score = cog_sim * 0.5 + per_sim * 0.3 + disc_weight * 0.2

        results.append({
            "discipline_name": disc_name,
            "cognitive_sim": round(cog_sim, 4),
            "persona_sim": round(per_sim, 4),
            "weight_bonus": round(disc_weight, 4),
            "score": round(score, 4),
        })

    results.sort(key=lambda x: x["score"], reverse=True)
    return results


# ============================================================================
# Layer 2: 专业类精选（Top 8 硬截断）
# ============================================================================

def layer2_category_match(
    user: UserProfile,
    data: FunnelData,
    l1_results: list[dict],
) -> list[dict]:
    """
    第二层：专业类精选
    匹配产业向往 + 资产敏感度 + 分数敏感度 → Top 8 截断

    score = industry_match × 0.5 + asset_match × 0.2 + score_match × 0.3
    """
    # 用户顶层产业（得分 > 60 的集群）
    top_industries = {
        ind for ind, score in user.macro_industry_vector.items()
        if score >= 60
    }
    if not top_industries:
        top_industries = {"互联网与软件"}  # 默认

    rank_tier = _rank_tier(user.estimated_rank_percentile)
    econ_level = user.family_economic_level  # "高"/"中"/"低"
    risk_tol = user.macro_value_vector.get("风险容忍度", 50.0)
    risk_tier = _risk_tier(risk_tol)

    # 构建 L1 的门类得分查找表
    disc_scores = {d["discipline_name"]: d["score"] for d in l1_results}

    results = []
    for cat_name, cat_labels in data.categories.items():
        disc_name = data.get_discipline(cat_name)

        # --- industry_match ---
        industry_tags = cat_labels.get("industry_map", [])
        matched_clusters = set()
        for tag in industry_tags:
            cluster = INDUSTRY_TAG_TO_CLUSTER.get(tag, "")
            if cluster in top_industries:
                matched_clusters.add(cluster)
        industry_match = len(matched_clusters) / max(len(top_industries), 1)
        # 奖励更广泛的产业覆盖
        industry_match = min(1.0, industry_match * (1.0 + 0.1 * len(industry_tags)))

        # --- asset_match ---
        asset_sens = cat_labels.get("asset_sensitivity", "中")
        asset_match = ASSET_ALIGNMENT_MATRIX.get(econ_level, {}).get(asset_sens, 0.5)

        # --- score_match ---
        score_sens = cat_labels.get("score_sensitivity", "中")
        score_match = SCORE_ALIGNMENT_MATRIX.get(rank_tier, {}).get(score_sens, 0.5)

        # 综合得分
        score = industry_match * 0.5 + asset_match * 0.2 + score_match * 0.3

        # 产业热度微调系数（0.95 ~ 1.05）
        industry_bonus = 1.0
        for tag in industry_tags:
            if tag in INDUSTRY_HEAT_BONUS:
                industry_bonus = max(industry_bonus, INDUSTRY_HEAT_BONUS[tag])
        score = score * industry_bonus

        # L1 门类得分微调（±5%）
        disc_bonus = disc_scores.get(disc_name, 0.5)
        score = score * (0.95 + 0.1 * disc_bonus)

        results.append({
            "category_name": cat_name,
            "discipline_name": disc_name,
            "industry_match": round(industry_match, 4),
            "asset_match": round(asset_match, 4),
            "score_match": round(score_match, 4),
            "score": round(min(1.0, score), 4),
            "labels": cat_labels,
        })

    results.sort(key=lambda x: x["score"], reverse=True)

    # 🔴 硬截断：仅保留 Top 8
    top8 = results[:8]

    return top8


# ============================================================================
# Layer 3: 专业微观狙击（≤6/类 硬截断）
# ============================================================================

def layer3_major_match(
    user: UserProfile,
    data: FunnelData,
    top_categories: list[dict],
) -> list[dict]:
    """
    第三层：专业微观狙击
    在 Top 8 专业类下计算每个专业的微观匹配，每类最多 6 个

    score = (micro_match × 0.6 + heat_align × 0.4) × threshold_pass
    """
    user_behavior = [user.micro_behavior_vector.get(d, 50.0) / 100.0 for d in BEHAVIOR_DIMENSIONS]
    user_physical = set(user.physical_conditions)
    risk_tol = user.macro_value_vector.get("风险容忍度", 50.0)
    risk_tier = _risk_tier(risk_tol)

    funnel_output = []

    for cat in top_categories:
        cat_name = cat["category_name"]
        disc_name = cat["discipline_name"]
        majors = data.get_majors_in_category(cat_name)

        major_results = []
        for major in majors:
            # --- micro_match ---
            micro_tags = major.get("micro_actions", [])
            major_behavior_vec = tags_to_vector(micro_tags, MICRO_ACTION_TO_BEHAVIOR, BEHAVIOR_DIMENSIONS)
            micro_match = cosine_similarity(user_behavior, major_behavior_vec)

            # --- heat_align ---
            social_heat = major.get("social_heat", "中")
            heat_align = HEAT_ALIGNMENT_MATRIX.get(risk_tier, {}).get(social_heat, 0.5)

            # --- threshold_pass ---
            hard_thresholds = major.get("hard_threshold", [])
            threshold_pass = 1
            triggered = []
            if hard_thresholds:
                for ht in hard_thresholds:
                    for uc in user_physical:
                        # 子串匹配：用户的"色盲"匹配专业的"无红绿色盲"
                        if uc in ht or ht in uc:
                            threshold_pass = 0
                            triggered.append(ht)
                            break

            # 🔴 最终公式（含招生体量市场容量系数）
            enroll_vol = major.get("enrollment_volume", "中")
            capacity_coef = ENROLLMENT_CAPACITY_COEFFICIENT.get(enroll_vol, 1.0)
            score = (micro_match * 0.6 + heat_align * 0.4) * capacity_coef * threshold_pass

            major_results.append({
                "major_name": major["name"],
                "major_code": major["code"],
                "category_name": cat_name,
                "discipline_name": disc_name,
                "micro_match": round(micro_match, 4),
                "heat_align": round(heat_align, 4),
                "threshold_pass": bool(threshold_pass),
                "triggered_thresholds": triggered,
                "social_heat": social_heat,
                "heat_trend": major.get("heat_trend", "平稳"),
                "hard_threshold": hard_thresholds,
                "micro_actions": micro_tags,
                "score": round(score, 4),
            })

        # 过滤 threshold_pass=0 的专业 + 排序
        major_results = [m for m in major_results if m["threshold_pass"]]
        major_results.sort(key=lambda x: x["score"], reverse=True)

        # 🔴 硬截断：每类最多 6 个
        top_majors = major_results[:6]

        if top_majors:
            funnel_output.append({
                "category_name": cat_name,
                "discipline_name": disc_name,
                "category_score": cat["score"],
                "category_labels": cat["labels"],
                "recommended_majors": top_majors,
            })

    return funnel_output


# ============================================================================
# 标签与理由生成
# ============================================================================

HOT_INDUSTRY_TAGS = {
    "人工智能/大模型", "半导体/集成电路", "新能源汽车",
    "智能制造/机器人", "新能源", "碳中和",
    "互联网/软件", "航空航天",
}


def _generate_major_tags(major: dict, cat_labels: dict) -> list[str]:
    """为专业生成展示标签"""
    tags = []
    micro = major.get("micro_match", 0)

    if micro >= 0.85:
        tags.append("[微观极度契合]")
    elif micro >= 0.70:
        tags.append("[微观高度契合]")

    score_sens = cat_labels.get("score_sensitivity", "中")
    if score_sens in ("极高", "高"):
        tags.append("[高分敏感]")

    asset_sens = cat_labels.get("asset_sensitivity", "中")
    if asset_sens == "低":
        tags.append("[低资源友好]")

    industry_tags = set(cat_labels.get("industry_map", []))
    if industry_tags & HOT_INDUSTRY_TAGS:
        tags.append("[产业风口]")

    heat = major.get("social_heat", "中")
    if heat == "极高":
        tags.append("[极度内卷]")

    if major.get("heat_trend") == "上升":
        tags.append("[热度上升]")

    return tags


def _generate_category_reason(cat: dict, user: UserProfile) -> str:
    """生成专业类级别的推荐理由"""
    parts = []

    # 产业匹配
    ind_match = cat.get("industry_match", 0)
    if ind_match >= 0.7:
        top_inds = [
            ind for ind, score in user.macro_industry_vector.items()
            if score >= 70
        ]
        if top_inds:
            parts.append(f"你强烈向往的{'/'.join(top_inds[:3])}赛道与此方向高度对口")

    # 资产匹配
    asset_match = cat.get("asset_match", 0)
    cat_labels = cat.get("category_labels", {})
    asset_sens = cat_labels.get("asset_sensitivity", "中")
    if asset_match >= 0.8 and asset_sens == "低":
        parts.append("技术驱动为主，家庭资源门槛低")
    elif asset_match < 0.4:
        parts.append(f"该方向对家庭资源有一定要求（资产敏感度：{asset_sens}）")

    # 分数匹配
    score_match = cat.get("score_match", 0)
    score_sens = cat_labels.get("score_sensitivity", "中")
    if score_match >= 0.8:
        parts.append(f"你的分数位次与此方向的竞争格局匹配良好")
    elif score_match < 0.4:
        parts.append(f"注意：此方向分数竞争激烈（敏感度：{score_sens}）")

    if not parts:
        parts.append("综合多维度评估后的推荐方向")

    return "；".join(parts)


def _generate_major_reason(major: dict, user: UserProfile) -> str:
    """生成具体专业的推荐理由"""
    parts = []
    micro = major.get("micro_match", 0)
    actions = major.get("micro_actions", [])

    if micro >= 0.85:
        top_actions = actions[:3] if actions else []
        action_str = "、".join(top_actions)
        parts.append(f"你的微观行为模式（{action_str}）与此专业核心要求极度吻合")

    elif micro >= 0.70:
        top_actions = actions[:2] if actions else []
        action_str = "、".join(top_actions)
        parts.append(f"你在{action_str}等方面与此专业有较好匹配")

    else:
        top_behaviors = user.get_top_behaviors(2)
        bh_str = "、".join(f"{k}({v:.0f}分)" for k, v in top_behaviors)
        parts.append(f"你的核心优势（{bh_str}）与此专业有一定关联")

    # 热度提示
    heat = major.get("social_heat", "中")
    if heat == "极高":
        parts.append("该专业当前竞争极为激烈，建议做好梯度规划")
    elif heat == "低":
        parts.append("该专业相对冷门但稳定，竞争压力较小")

    return "。".join(parts)


# ============================================================================
# 主入口：运行三层漏斗
# ============================================================================

def run_funnel(user: UserProfile, verbose: bool = False) -> list[dict]:
    """
    运行三层递进漏斗，返回结构化推荐结果。

    Args:
        user: 用户画像
        verbose: 是否打印中间过程

    Returns:
        list[dict]: 符合前端 Schema 的推荐结果
    """
    try:
        return _run_funnel_impl(user, verbose)
    except Exception as e:
        import traceback
        print(f"[ERROR] run_funnel failed: {e}")
        traceback.print_exc()
        return []


def _run_funnel_impl(user: UserProfile, verbose: bool = False) -> list[dict]:
    data = load_funnel_data()

    # 数据校验
    if not data.disciplines:
        print("[ERROR] Funnel data: no disciplines loaded!")
        return []
    if not data.categories:
        print("[ERROR] Funnel data: no categories loaded!")
        return []
    if not data.majors:
        print("[ERROR] Funnel data: no majors loaded!")
        return []

    # ---- Layer 1: 学科门类初筛 ----
    if verbose:
        print("\n[Layer 1] 学科门类初筛")
    l1 = layer1_discipline_match(user, data)
    if verbose:
        for d in l1[:5]:
            print(f"  {d['discipline_name']}: {d['score']:.4f} "
                  f"(cog={d['cognitive_sim']:.3f} per={d['persona_sim']:.3f} w={d['weight_bonus']:.2f})")

    # ---- Layer 2: 专业类精选 (Top 8) ----
    if verbose:
        print(f"\n[Layer 2] 专业类精选 → Top 8 截断")
    l2 = layer2_category_match(user, data, l1)
    if verbose:
        for i, c in enumerate(l2):
            print(f"  #{i+1} {c['category_name']} ({c['discipline_name']}): {c['score']:.4f} "
                  f"ind={c['industry_match']:.3f} ast={c['asset_match']:.3f} scr={c['score_match']:.3f}")

    if not l2:
        print("[ERROR] Layer 2 returned no categories!")
        return []

    # ---- Layer 3: 专业微观狙击 (≤6/类) ----
    if verbose:
        print(f"\n[Layer 3] 专业微观狙击 → ≤6/类 截断")
    l3 = layer3_major_match(user, data, l2)
    if verbose:
        for cat in l3:
            print(f"  [{cat['category_name']}] {len(cat['recommended_majors'])} majors")
            for m in cat["recommended_majors"][:3]:
                print(f"    {m['major_name']}: {m['score']:.4f} "
                      f"(micro={m['micro_match']:.3f} heat={m['heat_align']:.3f})")

    # ---- 格式化输出 ----
    output = []
    for cat in l3:
        cat_labels = cat.get("category_labels", {})
        majors_out = []
        for m in cat["recommended_majors"]:
            tags = _generate_major_tags(m, cat_labels)
            reason = _generate_major_reason(m, user)
            majors_out.append({
                "major_name": m["major_name"],
                "major_code": m["major_code"],
                "tags": tags,
                "major_reason": reason,
                "major_score": m["score"],
                "threshold_pass": m["threshold_pass"],
            })

        cat_reason = _generate_category_reason(cat, user)
        output.append({
            "category_name": cat["category_name"],
            "discipline_name": cat["discipline_name"],
            "category_reason": cat_reason,
            "category_score": cat["category_score"],
            "recommended_majors": majors_out,
        })

    return output


# ============================================================================
# 便捷打印
# ============================================================================

def print_funnel_results(results: list[dict]) -> None:
    """美化打印漏斗结果"""
    print()
    print("=" * 72)
    print("  🎯 三层递进漏斗推荐引擎 v6.0 — 最终推荐结果")
    print("=" * 72)

    for ci, cat in enumerate(results):
        medal = {0: "🥇", 1: "🥈", 2: "🥉"}.get(ci, f"#{ci+1}")
        print(f"\n{medal} [{cat['category_name']}] ({cat['discipline_name']})")
        print(f"   得分: {cat['category_score']:.3f}")
        print(f"   📌 {cat['category_reason']}")
        print(f"   ──────────────────────────────")

        for mi, m in enumerate(cat["recommended_majors"]):
            tag_str = " ".join(m["tags"]) if m["tags"] else ""
            print(f"   {mi+1}. {m['major_name']} ({m['major_code']}) "
                  f"| {m['major_score']:.3f} {tag_str}")
            print(f"      {m['major_reason']}")

    print(f"\n{'='*72}")
    total_majors = sum(len(c["recommended_majors"]) for c in results)
    print(f"  共 {len(results)} 个专业类 / {total_majors} 个专业")


# ============================================================================
# 自测
# ============================================================================

if __name__ == "__main__":
    from user_profile import create_test_user

    user = create_test_user()
    # 补充价值观向量（测试用户默认没有）
    if not user.macro_value_vector:
        user.macro_value_vector = {
            "稳定偏好": 30.0,
            "成长导向": 85.0,
            "风险容忍度": 60.0,
            "社会影响力": 45.0,
            "经济回报": 80.0,
        }

    results = run_funnel(user, verbose=True)
    print_funnel_results(results)

    # 验证
    print("\n[验证]")
    assert len(results) <= 8, f"Too many categories: {len(results)}"
    for cat in results:
        assert len(cat["recommended_majors"]) <= 6, \
            f"{cat['category_name']} has {len(cat['recommended_majors'])} majors"
    print(f"✅ 类别数: {len(results)} (≤8)")
    for cat in results:
        print(f"✅ {cat['category_name']}: {len(cat['recommended_majors'])} majors (≤6)")
    print("✅ 所有验证通过！")
